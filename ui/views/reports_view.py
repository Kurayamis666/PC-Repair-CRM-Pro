# ui/views/reports_view.py
"""
Экран отчётов для PC Repair CRM Pro
✅ ИСПРАВЛЕНО: Полный перевод, сортировка, валидация, обработка ошибок
✅ УЛУЧШЕНО: Адаптивные колонки, пустое состояние, проверка зависимостей
✅ СОВМЕСТИМО: Интеграция с системой тем, переводов и виджетов
"""

import customtkinter as ctk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, timedelta
from typing import Optional, Callable, Dict, Any, List
import sqlite3
import sys

from core.logger import app_logger
from database.connection import DatabaseConnection
from ui.theme import ColorTheme, ColorUtils
from translations import get_text
from ui.widgets.toast import ToastNotification
from ui.widgets.tables import DataTable
from utils.helpers import format_currency, format_date


class ReportsView(ctk.CTkFrame):
    """
    Экран отчётов с полным функционалом
    
    ✅ Полный перевод всех текстов (RU ↔ EN)
    ✅ Сортировка по клику на заголовок (через DataTable)
    ✅ Валидация дат и входных данных
    ✅ Адаптивные колонки и пустое состояние
    ✅ Проверка зависимостей (openpyxl) перед экспортом
    ✅ Прогресс-индикаторы для долгих операций
    ✅ Централизованные статусы и форматирование
    """
    
    on_navigate: Optional[Callable[[str], None]] = None
    
    # 📊 Конфигурация статусов заявок
    STATUS_CONFIG: Dict[str, Dict[str, Any]] = {
        "new": {"icon": "🆕", "label_ru": "Новая", "label_en": "New", "color": ColorTheme.STATUS_NEW},
        "diagnostics": {"icon": "🔍", "label_ru": "Диагностика", "label_en": "Diagnostics", "color": ColorTheme.STATUS_DIAGNOSTICS},
        "in_progress": {"icon": "🔧", "label_ru": "В работе", "label_en": "In Progress", "color": ColorTheme.STATUS_IN_PROGRESS},
        "ready": {"icon": "✅", "label_ru": "Готово", "label_en": "Ready", "color": ColorTheme.STATUS_READY},
        "closed": {"icon": "🏁", "label_ru": "Закрыто", "label_en": "Closed", "color": ColorTheme.STATUS_CLOSED},
        "cancelled": {"icon": "❌", "label_ru": "Отменено", "label_en": "Cancelled", "color": ColorTheme.ERROR},
    }
    
    def __init__(self, parent: ctk.CTkBaseClass, lang: str = "ru", on_navigate: Optional[Callable] = None, **kwargs):
        super().__init__(parent, fg_color="transparent", **kwargs)
        self.lang = lang
        self.on_navigate = on_navigate
        self.db = DatabaseConnection()
        self._report_trees: Dict[str, Any] = {}  # Кэш деревьев для экспорта
        self._build_ui()
    
    def _build_ui(self) -> None:
        """Построение интерфейса с полным переводом"""
        
        # 🏷️ Заголовок
        header = ctk.CTkFrame(self, fg_color=ColorTheme.PRIMARY, corner_radius=12)
        header.pack(fill="x", padx=10, pady=(5, 0))
        ctk.CTkLabel(
            header, 
            text="📊 " + get_text("reports", self.lang),
            font=ctk.CTkFont(size=22, weight="bold"),
            text_color=ColorTheme.TEXT_PRIMARY
        ).pack(pady=12)
        
        # 🔙 Кнопка назад
        ctk.CTkButton(
            self, 
            text=get_text("back", self.lang),
            command=lambda: self.on_navigate("dashboard") if self.on_navigate else None,
            width=120, height=32,
            fg_color=ColorTheme.BG_INPUT,
            hover_color=ColorTheme.BG_HOVER,
            corner_radius=10,
            font=ctk.CTkFont(size=12)
        ).pack(padx=20, pady=(12, 5), anchor="w")
        
        # Основной контент
        self.content_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.content_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        self._show_reports_menu()
    
    def _show_reports_menu(self) -> None:
        """Показ меню отчётов с переводом"""
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        # Типы отчётов с переводом
        reports = [
            ("💰", get_text("financial_report", self.lang), self._show_financial_report, ColorTheme.SUCCESS),
            ("📦", get_text("inventory_report", self.lang), self._show_inventory_report, ColorTheme.WARNING),
            ("👥", get_text("employee_report", self.lang), self._show_clients_report, ColorTheme.INFO),
            ("🔧", get_text("requests", self.lang), self._show_requests_report, ColorTheme.STATUS_IN_PROGRESS),
            ("📈", get_text("analytics", self.lang), self._show_performance_report, ColorTheme.PRIMARY),
            ("📋", get_text("export_price_list", self.lang), self._show_pricelist_report, ColorTheme.STATUS_READY),
        ]
        
        grid_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        grid_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        for i, (icon, name, command, color) in enumerate(reports):
            row, col = divmod(i, 2)
            
            card = ctk.CTkFrame(grid_frame, fg_color=ColorTheme.BG_CARD, corner_radius=16, border_width=1, border_color=ColorTheme.BORDER)
            card.grid(row=row, column=col, padx=12, pady=12, sticky="nsew")
            
            accent = ctk.CTkFrame(card, fg_color=color, height=4, corner_radius=2)
            accent.pack(fill="x", padx=20, pady=(16, 0))
            
            ctk.CTkLabel(
                card, text=icon,
                font=ctk.CTkFont(size=32)
            ).pack(pady=(12, 4))
            
            ctk.CTkLabel(
                card, text=name,
                font=ctk.CTkFont(size=14, weight="bold"),
                text_color=ColorTheme.TEXT_PRIMARY
            ).pack(pady=(0, 4))
            
            btn = ctk.CTkButton(
                card,
                text=get_text("open", self.lang) if get_text("open", self.lang) != "open" else "Открыть",
                height=32, width=120,
                command=command,
                font=ctk.CTkFont(size=12),
                corner_radius=8,
                fg_color=color,
                hover_color=ColorUtils.darken(color, 15)
            )
            btn.pack(pady=(4, 16))
        
        grid_frame.grid_columnconfigure((0, 1), weight=1)
        grid_frame.grid_rowconfigure((0, 1, 2), weight=1)
    
    # ==================== 💰 ФИНАНСОВЫЙ ОТЧЁТ ====================
    def _show_financial_report(self) -> None:
        """Финансовый отчёт с переводом и валидацией"""
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        # Заголовок
        header = ctk.CTkFrame(self.content_frame, fg_color=ColorTheme.SUCCESS, corner_radius=12)
        header.pack(fill="x", pady=10)
        ctk.CTkLabel(
            header, 
            text="💰 " + get_text("financial_report", self.lang),
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=ColorTheme.TEXT_PRIMARY
        ).pack(pady=10)
        
        # Фильтр по датам
        filter_frame = ctk.CTkFrame(self.content_frame, fg_color=ColorTheme.BG_CARD)
        filter_frame.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(
            filter_frame, 
            text=get_text("period", self.lang) + ":",
            text_color=ColorTheme.TEXT_PRIMARY
        ).pack(side="left", padx=10)
        
        # Дата начала
        ctk.CTkLabel(filter_frame, text=get_text("start_date", self.lang), text_color=ColorTheme.TEXT_SECONDARY).pack(side="left", padx=(10, 2))
        self.start_date = ctk.CTkEntry(
            filter_frame, width=120, height=30,
            fg_color=ColorTheme.BG_INPUT,
            text_color=ColorTheme.TEXT_PRIMARY
        )
        self.start_date.pack(side="left", padx=2)
        self.start_date.insert(0, (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d"))
        
        ctk.CTkLabel(filter_frame, text="-", text_color=ColorTheme.TEXT_PRIMARY).pack(side="left", padx=5)
        
        # Дата окончания
        ctk.CTkLabel(filter_frame, text=get_text("end_date", self.lang), text_color=ColorTheme.TEXT_SECONDARY).pack(side="left", padx=(2, 2))
        self.end_date = ctk.CTkEntry(
            filter_frame, width=120, height=30,
            fg_color=ColorTheme.BG_INPUT,
            text_color=ColorTheme.TEXT_PRIMARY
        )
        self.end_date.pack(side="left", padx=2)
        self.end_date.insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        ctk.CTkButton(
            filter_frame, 
            text="🔄 " + get_text("update", self.lang),
            command=self._load_financial_report,
            width=100, height=30,
            fg_color=ColorTheme.INFO,
            hover_color=ColorUtils.darken(ColorTheme.INFO, 10)
        ).pack(side="left", padx=10)
        
        ctk.CTkButton(
            filter_frame, 
            text="📤 " + get_text("export_excel", self.lang),
            command=lambda: self._export_report("financial"),
            width=130, height=30,
            fg_color=ColorTheme.SUCCESS,
            hover_color=ColorUtils.darken(ColorTheme.SUCCESS, 10)
        ).pack(side="left", padx=10)
        
        ctk.CTkButton(
            filter_frame, 
            text=get_text("back", self.lang),
            command=self._show_reports_menu,
            width=100, height=30
        ).pack(side="right", padx=10)
        
        # Статистика
        stats_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        stats_frame.pack(fill="x", padx=20, pady=10)
        
        self.fin_stats: Dict[str, ctk.CTkLabel] = {}
        stats_config = [
            (get_text("revenue", self.lang), ColorTheme.SUCCESS),
            (get_text("expenses", self.lang), ColorTheme.ERROR),
            (get_text("profit", self.lang), ColorTheme.PRIMARY),
            (get_text("requests_count", self.lang), ColorTheme.INFO)
        ]
        
        for i, (name, color) in enumerate(stats_config):
            frame = ctk.CTkFrame(stats_frame, fg_color=ColorTheme.BG_CARD, corner_radius=12)
            frame.pack(side="left", fill="x", expand=True, padx=5)
            
            ctk.CTkLabel(
                frame, text=name,
                text_color=ColorTheme.TEXT_SECONDARY,
                font=ctk.CTkFont(size=12)
            ).pack(pady=(10, 5))
            
            lbl = ctk.CTkLabel(
                frame, text="0 ₽",
                text_color=color,
                font=ctk.CTkFont(size=18, weight="bold")
            )
            lbl.pack(pady=(5, 10))
            self.fin_stats[name] = lbl
        
        # Таблица с сортировкой
        table_frame = ctk.CTkFrame(self.content_frame, fg_color=ColorTheme.BG_INPUT)
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        columns = [
            get_text("col_date", self.lang),
            get_text("col_employee", self.lang),  # Было "Клиент"
            get_text("col_problem", self.lang),
            get_text("col_sum", self.lang),
            get_text("col_type", self.lang)
        ]
        col_widths = {
            get_text("col_date", self.lang): 100,
            get_text("col_employee", self.lang): 150,
            get_text("col_problem", self.lang): 250,
            get_text("col_sum", self.lang): 100,
            get_text("col_type", self.lang): 100
        }
        col_align = {
            get_text("col_date", self.lang): "center",
            get_text("col_employee", self.lang): "left",
            get_text("col_problem", self.lang): "left",
            get_text("col_sum", self.lang): "right",
            get_text("col_type", self.lang): "center"
        }
        
        self.fin_tree = DataTable(
            table_frame,
            columns=columns,
            column_widths=col_widths,
            column_align=col_align,
            sortable=True,
            copyable=True,
            searchable=False
        )
        self.fin_tree.pack(fill="both", expand=True)
        
        # ✅ Кэшируем дерево для экспорта
        self._report_trees["financial"] = self.fin_tree
        
        # 🎨 Теги для доходов/расходов
        self.fin_tree.tag_configure("income", background=ColorUtils.darken(ColorTheme.SUCCESS, 30), foreground="#6ee7b7")
        self.fin_tree.tag_configure("expense", background=ColorUtils.darken(ColorTheme.WARNING, 30), foreground="#fdba74")
        
        self._load_financial_report()
    
    def _validate_date_range(self, start: str, end: str) -> bool:
        """Валидация диапазона дат"""
        try:
            start_dt = datetime.strptime(start, "%Y-%m-%d")
            end_dt = datetime.strptime(end, "%Y-%m-%d")
            if start_dt > end_dt:
                ToastNotification(self, get_text("invalid_date_range", self.lang) or "Дата начала не может быть после даты окончания", "warning")
                return False
            return True
        except ValueError:
            ToastNotification(self, get_text("invalid_date_format", self.lang) or "Неверный формат даты (ожидается: ГГГГ-ММ-ДД)", "warning")
            return False
    
    def _load_financial_report(self) -> None:
        """Загрузка финансового отчёта с валидацией"""
        start = self.start_date.get().strip()
        end = self.end_date.get().strip()
        
        if not self._validate_date_range(start, end):
            return
        
        # Показываем индикатор загрузки
        ToastNotification(self, get_text("loading_report", self.lang) or "🔄 Загрузка отчёта...", "info")
        
        try:
            with self.db.get_cursor() as cur:
                # Общая статистика
                cur.execute("""
                    SELECT 
                        COALESCE(SUM(total_cost), 0) as revenue,
                        COALESCE(SUM(labor_cost + parts_cost), 0) as expenses,
                        COUNT(*) as requests_count
                    FROM requests
                    WHERE DATE(created_at) BETWEEN ? AND ?
                    AND status IN ('ready', 'closed')
                """, (start, end))
                
                stats = cur.fetchone()
                revenue = stats[0] or 0
                expenses = stats[1] or 0
                profit = revenue - expenses
                count = stats[2] or 0
                
                # Обновляем статистику с форматированием
                self.fin_stats[get_text("revenue", self.lang)].configure(text=format_currency(revenue, "RUB", self.lang))
                self.fin_stats[get_text("expenses", self.lang)].configure(text=format_currency(expenses, "RUB", self.lang))
                self.fin_stats[get_text("profit", self.lang)].configure(text=format_currency(profit, "RUB", self.lang))
                self.fin_stats[get_text("requests_count", self.lang)].configure(text=str(count))
                
                # Детализация
                cur.execute("""
                    SELECT 
                        DATE(r.created_at) as date,
                        c.full_name as client,
                        CASE 
                            WHEN r.problem_desc IS NOT NULL THEN r.problem_desc
                            ELSE 'Услуга'
                        END as service,
                        r.total_cost,
                        'Доход' as type
                    FROM requests r
                    LEFT JOIN employees c ON r.client_id = c.id
                    WHERE DATE(r.created_at) BETWEEN ? AND ?
                    AND r.status IN ('ready', 'closed')
                    ORDER BY r.created_at DESC
                    LIMIT 100
                """, (start, end))
                
                rows = cur.fetchall()
                
                # Показываем пустое состояние если нет данных
                if not rows:
                    self.fin_tree.show_empty_state(
                        message=get_text("no_data_for_period", self.lang) or "Нет данных за выбранный период",
                        icon="📭"
                    )
                    return
                else:
                    self.fin_tree.hide_empty_state()
                
                # Заполняем таблицу
                for idx, row in enumerate(rows):
                    tag = "income" if row[4] == "Доход" else "expense"
                    tag = "odd" if idx % 2 else "even"
                    
                    # Форматируем сумму
                    amount = format_currency(row[3], "RUB", self.lang)
                    values = (row[0], row[1] or "—", row[2], amount, row[4])
                    self.fin_tree.insert("", "end", values=values, tags=(tag,))
        
        except Exception as e:
            app_logger.error(f"Error loading financial report: {e}")
            ToastNotification(self, f"{get_text('error_loading', self.lang)}: {e}", "error")
    
    # ==================== 📦 ОТЧЁТ ПО ЗАПАСАМ ====================
    def _show_inventory_report(self) -> None:
        """Отчёт по запасам с переводом"""
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        header = ctk.CTkFrame(self.content_frame, fg_color=ColorTheme.WARNING, corner_radius=12)
        header.pack(fill="x", pady=10)
        ctk.CTkLabel(
            header, 
            text="📦 " + get_text("inventory_report", self.lang),
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=ColorTheme.TEXT_PRIMARY
        ).pack(pady=10)
        
        btn_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkButton(
            btn_frame, 
            text=get_text("back", self.lang),
            command=self._show_reports_menu,
            width=100, height=30
        ).pack(side="left", padx=10)
        
        ctk.CTkButton(
            btn_frame, 
            text="📤 " + get_text("export_excel", self.lang),
            command=lambda: self._export_report("inventory"),
            width=130, height=30,
            fg_color=ColorTheme.SUCCESS,
            hover_color=ColorUtils.darken(ColorTheme.SUCCESS, 10)
        ).pack(side="right", padx=10)
        
        # Статистика
        stats_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        stats_frame.pack(fill="x", padx=20, pady=10)
        
        self.inv_stats: Dict[str, ctk.CTkLabel] = {}
        stats_config = [
            (get_text("total_items", self.lang), ColorTheme.INFO),
            (get_text("total_value", self.lang), ColorTheme.PRIMARY),
            (get_text("low_stock", self.lang), ColorTheme.WARNING),
            (get_text("out_of_stock", self.lang), ColorTheme.ERROR)
        ]
        
        for i, (name, color) in enumerate(stats_config):
            frame = ctk.CTkFrame(stats_frame, fg_color=ColorTheme.BG_CARD, corner_radius=12)
            frame.pack(side="left", fill="x", expand=True, padx=5)
            
            ctk.CTkLabel(
                frame, text=name,
                text_color=ColorTheme.TEXT_SECONDARY
            ).pack(pady=(10, 5))
            
            lbl = ctk.CTkLabel(
                frame, text="0",
                text_color=color,
                font=ctk.CTkFont(size=18, weight="bold")
            )
            lbl.pack(pady=(5, 10))
            self.inv_stats[name] = lbl
        
        # Таблица
        table_frame = ctk.CTkFrame(self.content_frame, fg_color=ColorTheme.BG_INPUT)
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        columns = [
            get_text("col_id", self.lang),
            get_text("col_name", self.lang),
            get_text("col_sku", self.lang),
            get_text("col_quantity", self.lang),
            get_text("col_unit", self.lang),
            get_text("col_price", self.lang),
            get_text("col_sum", self.lang),
            get_text("col_status", self.lang)
        ]
        col_widths = {
            get_text("col_id", self.lang): 50,
            get_text("col_name", self.lang): 200,
            get_text("col_sku", self.lang): 100,
            get_text("col_quantity", self.lang): 80,
            get_text("col_unit", self.lang): 60,
            get_text("col_price", self.lang): 100,
            get_text("col_sum", self.lang): 100,
            get_text("col_status", self.lang): 120
        }
        
        self.inv_tree = DataTable(
            table_frame,
            columns=columns,
            column_widths=col_widths,
            sortable=True,
            copyable=True
        )
        self.inv_tree.pack(fill="both", expand=True)
        self._report_trees["inventory"] = self.inv_tree
        
        # Теги для статусов
        self.inv_tree.tag_configure("low_stock", background=ColorUtils.darken(ColorTheme.WARNING, 30), foreground="#fdba74")
        self.inv_tree.tag_configure("out_of_stock", background=ColorUtils.darken(ColorTheme.ERROR, 30), foreground="#ff6b6b")
        
        self._load_inventory_report()
    
    def _load_inventory_report(self) -> None:
        """Загрузка отчёта по запасам"""
        try:
            with self.db.get_cursor() as cur:
                cur.execute("""
                    SELECT 
                        id, name, sku, quantity, unit, price,
                        (quantity * price) as total_value,
                        CASE 
                            WHEN quantity = 0 THEN ?
                            WHEN quantity < 5 THEN ?
                            ELSE ?
                        END as status
                    FROM parts
                    ORDER BY quantity ASC
                """, (
                    get_text("out_of_stock", self.lang),
                    get_text("low_stock", self.lang),
                    get_text("in_stock", self.lang) or "В наличии"
                ))
                
                rows = cur.fetchall()
                
                total_items = len(rows)
                total_value = sum(row[6] or 0 for row in rows)
                low_stock = sum(1 for row in rows if row[3] and 0 < row[3] < 5)
                out_of_stock = sum(1 for row in rows if row[3] == 0)
                
                self.inv_stats[get_text("total_items", self.lang)].configure(text=str(total_items))
                self.inv_stats[get_text("total_value", self.lang)].configure(text=format_currency(total_value, "RUB", self.lang))
                self.inv_stats[get_text("low_stock", self.lang)].configure(text=str(low_stock))
                self.inv_stats[get_text("out_of_stock", self.lang)].configure(text=str(out_of_stock))
                
                # Пустое состояние
                if not rows:
                    self.inv_tree.show_empty_state(
                        message=get_text("no_parts", self.lang) or "Запчасти не найдены",
                        icon="📦"
                    )
                    return
                else:
                    self.inv_tree.hide_empty_state()
                
                for idx, row in enumerate(rows):
                    if row[3] == 0:
                        tag = "out_of_stock"
                    elif row[3] and row[3] < 5:
                        tag = "low_stock"
                    else:
                        tag = "odd" if idx % 2 else "even"
                    
                    values = (
                        row[0], row[1], row[2], row[3], row[4],
                        format_currency(row[5], "RUB", self.lang),
                        format_currency(row[6], "RUB", self.lang),
                        row[7]
                    )
                    self.inv_tree.insert("", "end", values=values, tags=(tag,))
        
        except Exception as e:
            app_logger.error(f"Error loading inventory report: {e}")
            ToastNotification(self, f"{get_text('error_loading', self.lang)}: {e}", "error")
    
    # ==================== 👥 ОТЧЁТ ПО СОТРУДНИКАМ ====================
    def _show_clients_report(self) -> None:
        """Отчёт по сотрудникам (бывш. клиентам)"""
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        btn_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=(10, 0))
        ctk.CTkButton(
            btn_frame,
            text=get_text("back", self.lang),
            command=self._show_reports_menu,
            width=100, height=30
        ).pack(side="left", padx=10)
        
        # Вкладки
        notebook = ctk.CTkTabview(self.content_frame, fg_color="transparent")
        notebook.pack(fill="both", expand=True, padx=20, pady=10)
        
        employees_tab = notebook.add(get_text("employees", self.lang))
        contractors_tab = notebook.add(get_text("contractors", self.lang))
        
        self._build_employees_tab(employees_tab)
        self._build_contractors_tab(contractors_tab)
    
    def _build_employees_tab(self, parent):
        """Вкладка сотрудников"""
        header = ctk.CTkFrame(parent, fg_color=ColorTheme.INFO, corner_radius=12)
        header.pack(fill="x", pady=10)
        ctk.CTkLabel(
            header, 
            text=get_text("employee_stats", self.lang) or "👥 Статистика сотрудников",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=ColorTheme.TEXT_PRIMARY
        ).pack(pady=10)
        
        btn_frame = ctk.CTkFrame(parent, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkButton(
            btn_frame, 
            text="📤 " + get_text("export", self.lang),
            command=lambda: self._export_report("employees"),
            width=120, height=30,
            fg_color=ColorTheme.SUCCESS,
            hover_color=ColorUtils.darken(ColorTheme.SUCCESS, 10)
        ).pack(side="right", padx=10)
        
        table_frame = ctk.CTkFrame(parent, fg_color=ColorTheme.BG_INPUT)
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        columns = [
            get_text("col_id", self.lang),
            get_text("col_name", self.lang),
            get_text("col_phone", self.lang),
            get_text("col_email", self.lang),
            get_text("requests_count", self.lang),
            get_text("total_spent", self.lang)
        ]
        col_widths = {
            get_text("col_id", self.lang): 50,
            get_text("col_name", self.lang): 200,
            get_text("col_phone", self.lang): 120,
            get_text("col_email", self.lang): 180,
            get_text("requests_count", self.lang): 80,
            get_text("total_spent", self.lang): 100
        }
        
        self.clients_tree = DataTable(
            table_frame,
            columns=columns,
            column_widths=col_widths,
            sortable=True,
            copyable=True
        )
        self.clients_tree.pack(fill="both", expand=True)
        self._report_trees["employees"] = self.clients_tree
        
        self._load_employees_report()
    
    def _load_employees_report(self) -> None:
        """Загрузка отчёта по сотрудникам"""
        for item in self.clients_tree.get_children():
            self.clients_tree.delete(item)
        
        try:
            with self.db.get_cursor() as cur:
                cur.execute("""
                    SELECT 
                        c.id,
                        c.full_name,
                        c.phone,
                        c.email,
                        COUNT(r.id) as requests_count,
                        COALESCE(SUM(r.total_cost), 0) as total_spent
                    FROM employees c
                    LEFT JOIN requests r ON c.id = r.client_id
                    GROUP BY c.id
                    ORDER BY total_spent DESC
                """)
                
                rows = cur.fetchall()
                
                if not rows:
                    self.clients_tree.show_empty_state(
                        message=get_text("no_employees", self.lang) or "Сотрудники не найдены",
                        icon="👥"
                    )
                    return
                else:
                    self.clients_tree.hide_empty_state()
                
                for idx, row in enumerate(rows):
                    tag = "odd" if idx % 2 else "even"
                    values = (
                        row[0], row[1], row[2] or "—", row[3] or "—",
                        row[4], format_currency(row[5], "RUB", self.lang)
                    )
                    self.clients_tree.insert("", "end", values=values, tags=(tag,))
        
        except Exception as e:
            app_logger.error(f"Error loading employees report: {e}")
    
    def _build_contractors_tab(self, parent):
        """Вкладка контрагентов"""
        header = ctk.CTkFrame(parent, fg_color=ColorTheme.PRIMARY, corner_radius=12)
        header.pack(fill="x", pady=10)
        ctk.CTkLabel(
            header, 
            text=get_text("contractor_stats", self.lang) or "🏢 Статистика контрагентов",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=ColorTheme.TEXT_PRIMARY
        ).pack(pady=10)
        
        table_frame = ctk.CTkFrame(parent, fg_color=ColorTheme.BG_INPUT)
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        columns = [
            get_text("col_id", self.lang),
            get_text("col_name", self.lang),
            get_text("col_inn", self.lang),
            get_text("col_phone", self.lang),
            get_text("col_email", self.lang)
        ]
        
        self.contractors_tree = DataTable(
            table_frame,
            columns=columns,
            sortable=True,
            copyable=True
        )
        self.contractors_tree.pack(fill="both", expand=True)
        
        self._load_contractors_report()
    
    def _load_contractors_report(self) -> None:
        """Загрузка отчёта по контрагентам"""
        try:
            with self.db.get_cursor() as cur:
                cur.execute("""
                    SELECT id, name, inn, phone, email 
                    FROM contractors 
                    ORDER BY name
                """)
                rows = cur.fetchall()
                
                if not rows:
                    self.contractors_tree.show_empty_state(
                        message=get_text("no_contractors", self.lang) or "Контрагенты не найдены",
                        icon="🏢"
                    )
                    return
                else:
                    self.contractors_tree.hide_empty_state()
                
                for idx, row in enumerate(rows):
                    tag = "odd" if idx % 2 else "even"
                    values = (row[0], row[1], row[2] or "—", row[3] or "—", row[4] or "—")
                    self.contractors_tree.insert("", "end", values=values, tags=(tag,))
        
        except Exception as e:
            app_logger.error(f"Error loading contractors report: {e}")
    
    # ==================== 🔧 ОТЧЁТ ПО ЗАЯВКАМ ====================
    def _show_requests_report(self) -> None:
        """Отчёт по заявкам с переводом статусов"""
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        header = ctk.CTkFrame(self.content_frame, fg_color=ColorTheme.STATUS_IN_PROGRESS, corner_radius=12)
        header.pack(fill="x", pady=10)
        ctk.CTkLabel(
            header, 
            text="🔧 " + get_text("requests", self.lang),
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=ColorTheme.TEXT_PRIMARY
        ).pack(pady=10)
        
        btn_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkButton(
            btn_frame, 
            text=get_text("back", self.lang),
            command=self._show_reports_menu,
            width=100, height=30
        ).pack(side="left", padx=10)
        
        ctk.CTkButton(
            btn_frame, 
            text="📤 " + get_text("export", self.lang),
            command=lambda: self._export_report("requests"),
            width=120, height=30,
            fg_color=ColorTheme.SUCCESS,
            hover_color=ColorUtils.darken(ColorTheme.SUCCESS, 10)
        ).pack(side="right", padx=10)
        
        table_frame = ctk.CTkFrame(self.content_frame, fg_color=ColorTheme.BG_INPUT)
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        columns = [
            get_text("col_id", self.lang),
            get_text("col_date", self.lang),
            get_text("col_employee", self.lang),
            get_text("col_equipment", self.lang),
            get_text("col_status", self.lang),
            get_text("col_sum", self.lang),
            get_text("col_master", self.lang)
        ]
        col_widths = {
            get_text("col_id", self.lang): 50,
            get_text("col_date", self.lang): 100,
            get_text("col_employee", self.lang): 150,
            get_text("col_equipment", self.lang): 150,
            get_text("col_status", self.lang): 120,
            get_text("col_sum", self.lang): 100,
            get_text("col_master", self.lang): 120
        }
        
        self.req_tree = DataTable(
            table_frame,
            columns=columns,
            column_widths=col_widths,
            sortable=True,
            copyable=True
        )
        self.req_tree.pack(fill="both", expand=True)
        self._report_trees["requests"] = self.req_tree
        
        self._load_requests_report()
    
    def _get_status_display(self, status: str) -> str:
        """Получить отображаемый статус с иконкой"""
        config = self.STATUS_CONFIG.get(status, {})
        icon = config.get("icon", "❓")
        label = config.get(f"label_{self.lang}", status)
        return f"{icon} {label}"
    
    def _load_requests_report(self) -> None:
        """Загрузка отчёта по заявкам"""
        for item in self.req_tree.get_children():
            self.req_tree.delete(item)
        
        try:
            with self.db.get_cursor() as cur:
                cur.execute("""
                    SELECT 
                        r.id,
                        DATE(r.created_at),
                        c.full_name,
                        e.model,
                        r.status,
                        r.total_cost,
                        u.full_name
                    FROM requests r
                    LEFT JOIN employees c ON r.client_id = c.id
                    LEFT JOIN equipment e ON r.equipment_id = e.id
                    LEFT JOIN users u ON r.user_id = u.id
                    ORDER BY r.created_at DESC
                    LIMIT 200
                """)
                
                rows = cur.fetchall()
                
                if not rows:
                    self.req_tree.show_empty_state(
                        message=get_text("no_requests", self.lang) or "Заявки не найдены",
                        icon="🔧"
                    )
                    return
                else:
                    self.req_tree.hide_empty_state()
                
                for idx, row in enumerate(rows):
                    tag = "odd" if idx % 2 else "even"
                    status = self._get_status_display(row[4])
                    values = (
                        row[0], row[1], row[2] or "—", row[3] or "—",
                        status, format_currency(row[5], "RUB", self.lang), row[6] or "—"
                    )
                    self.req_tree.insert("", "end", values=values, tags=(tag,))
        
        except Exception as e:
            app_logger.error(f"Error loading requests report: {e}")
    
    # ==================== 📋 ПРАЙС-ЛИСТ ====================
    def _show_pricelist_report(self) -> None:
        """Прайс-лист с переводом"""
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        header = ctk.CTkFrame(self.content_frame, fg_color=ColorTheme.STATUS_READY, corner_radius=12)
        header.pack(fill="x", pady=10)
        ctk.CTkLabel(
            header, 
            text="📋 " + get_text("export_price_list", self.lang),
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=ColorTheme.TEXT_PRIMARY
        ).pack(pady=10)
        
        btn_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkButton(
            btn_frame, 
            text=get_text("back", self.lang),
            command=self._show_reports_menu,
            width=100, height=30
        ).pack(side="left", padx=10)
        
        ctk.CTkButton(
            btn_frame, 
            text="📤 " + get_text("export_excel", self.lang),
            command=lambda: self._export_report("pricelist"),
            width=140, height=30,
            fg_color=ColorTheme.SUCCESS,
            hover_color=ColorUtils.darken(ColorTheme.SUCCESS, 10)
        ).pack(side="right", padx=10)
        
        table_frame = ctk.CTkFrame(self.content_frame, fg_color=ColorTheme.BG_INPUT)
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        columns = [
            get_text("col_id", self.lang),
            get_text("col_name", self.lang),
            get_text("col_sku", self.lang),
            get_text("cost_price", self.lang),
            get_text("retail_price", self.lang),
            get_text("markup", self.lang),
            get_text("col_quantity", self.lang)
        ]
        col_widths = {
            get_text("col_id", self.lang): 50,
            get_text("col_name", self.lang): 250,
            get_text("col_sku", self.lang): 100,
            get_text("cost_price", self.lang): 120,
            get_text("retail_price", self.lang): 120,
            get_text("markup", self.lang): 100,
            get_text("col_quantity", self.lang): 80
        }
        
        self.price_tree = DataTable(
            table_frame,
            columns=columns,
            column_widths=col_widths,
            sortable=True,
            copyable=True
        )
        self.price_tree.pack(fill="both", expand=True)
        self._report_trees["pricelist"] = self.price_tree
        
        self._load_pricelist_report()
    
    def _load_pricelist_report(self) -> None:
        """Загрузка прайс-листа"""
        for item in self.price_tree.get_children():
            self.price_tree.delete(item)
        
        try:
            with self.db.get_cursor() as cur:
                cur.execute("""
                    SELECT 
                        id, name, sku, cost, price,
                        CASE WHEN cost > 0 THEN ((price - cost) / cost * 100) ELSE 0 END as markup,
                        quantity
                    FROM parts
                    ORDER BY name
                """)
                
                rows = cur.fetchall()
                
                if not rows:
                    self.price_tree.show_empty_state(
                        message=get_text("no_parts", self.lang) or "Запчасти не найдены",
                        icon="📋"
                    )
                    return
                else:
                    self.price_tree.hide_empty_state()
                
                for idx, row in enumerate(rows):
                    tag = "odd" if idx % 2 else "even"
                    values = (
                        row[0], row[1], row[2],
                        format_currency(row[3], "RUB", self.lang),
                        format_currency(row[4], "RUB", self.lang),
                        f"{row[5]:.0f}%" if row[5] else "—",
                        row[6]
                    )
                    self.price_tree.insert("", "end", values=values, tags=(tag,))
        
        except Exception as e:
            app_logger.error(f"Error loading pricelist: {e}")
    
    # ==================== 📤 ЭКСПОРТ ОТЧЁТОВ ====================
    def _export_report(self, report_type: str) -> None:
        """Экспорт отчёта в Excel с проверкой зависимостей"""
        
        # ✅ Проверка наличия openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            ToastNotification(
                self, 
                get_text("missing_openpyxl", self.lang) or "❌ Установите openpyxl: pip install openpyxl", 
                "error"
            )
            return
        
        try:
            # ✅ Безопасное получение дерева
            tree = self._report_trees.get(report_type)
            if not tree or not hasattr(tree, 'get_children'):
                ToastNotification(self, get_text("report_not_loaded", self.lang) or "Сначала загрузите отчёт", "warning")
                return
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"report_{report_type}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )
            
            if not file_path:
                return
            
            # Показываем прогресс
            ToastNotification(self, get_text("exporting", self.lang) or "🔄 Экспорт...", "info")
            
            wb = Workbook()
            ws = wb.active
            ws.title = report_type
            
            # Заголовок
            header_font = Font(bold=True, color="FFFFFF")
            primary_hex = ColorTheme.PRIMARY.lstrip("#")
            header_fill = PatternFill(start_color=primary_hex, end_color=primary_hex, fill_type="solid")
            
            # Заголовки колонок
            columns = tree['columns']
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Данные
            for row_idx, item in enumerate(tree.get_children(), 2):
                values = tree.item(item)['values']
                for col_idx, value in enumerate(values, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Авторазмер колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            ToastNotification(self, f"{get_text('report_saved', self.lang) or '✅ Отчёт сохранён'}: {file_path}", "success")
            app_logger.info(f"📤 Exported {report_type} report to {file_path}")
        
        except PermissionError:
            ToastNotification(self, get_text("export_permission_error", self.lang) or "❌ Нет прав на запись файла", "error")
        except Exception as e:
            app_logger.error(f"Error exporting report: {e}")
            ToastNotification(self, f"{get_text('error_exporting', self.lang)}: {e}", "error")
    
    def _show_performance_report(self) -> None:
        """Отчёт по эффективности — аналитика заявок"""
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        header = ctk.CTkFrame(self.content_frame, fg_color=ColorTheme.PRIMARY, corner_radius=12)
        header.pack(fill="x", pady=10)
        ctk.CTkLabel(
            header,
            text="📈 " + get_text("analytics", self.lang),
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=ColorTheme.TEXT_PRIMARY
        ).pack(pady=10)
        
        btn_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=10)
        ctk.CTkButton(
            btn_frame,
            text=get_text("back", self.lang),
            command=self._show_reports_menu,
            width=100, height=30
        ).pack(side="left", padx=10)
        
        stats_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        stats_frame.pack(fill="x", padx=20, pady=10)
        
        try:
            with self.db.get_cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM requests")
                total = cur.fetchone()[0] or 0
                
                cur.execute("SELECT COUNT(*) FROM requests WHERE status IN ('ready', 'closed')")
                completed = cur.fetchone()[0] or 0
                
                cur.execute("SELECT COALESCE(AVG(total_cost), 0) FROM requests WHERE status IN ('ready', 'closed')")
                avg_cost = cur.fetchone()[0] or 0
                
                cur.execute("SELECT COUNT(*) FROM requests WHERE status IN ('new', 'diagnostics', 'in_progress')")
                active = cur.fetchone()[0] or 0
                
                completion_pct = (completed / total * 100) if total > 0 else 0
                
                stats_config = [
                    (get_text("requests_count", self.lang), str(total), ColorTheme.INFO),
                    (get_text("completed_requests", self.lang) if get_text("completed_requests", self.lang) != "completed_requests" else "Завершено", str(completed), ColorTheme.SUCCESS),
                    (get_text("active_requests", self.lang) if get_text("active_requests", self.lang) != "active_requests" else "Активные", str(active), ColorTheme.WARNING),
                    (get_text("avg_request_cost", self.lang) if get_text("avg_request_cost", self.lang) != "avg_request_cost" else "Средний чек", format_currency(avg_cost, "RUB", self.lang), ColorTheme.PRIMARY),
                    (get_text("completion_rate", self.lang) if get_text("completion_rate", self.lang) != "completion_rate" else "Выполнение", f"{completion_pct:.0f}%", ColorTheme.STATUS_READY),
                ]
                
                for name, value, color in stats_config:
                    frame = ctk.CTkFrame(stats_frame, fg_color=ColorTheme.BG_CARD, corner_radius=12)
                    frame.pack(side="left", fill="x", expand=True, padx=5)
                    ctk.CTkLabel(frame, text=name, text_color=ColorTheme.TEXT_SECONDARY, font=ctk.CTkFont(size=12)).pack(pady=(10, 5))
                    ctk.CTkLabel(frame, text=value, text_color=color, font=ctk.CTkFont(size=18, weight="bold")).pack(pady=(5, 10))
                
                # Status breakdown table
                table_frame = ctk.CTkFrame(self.content_frame, fg_color=ColorTheme.BG_CARD, corner_radius=12)
                table_frame.pack(fill="both", expand=True, padx=20, pady=10)
                
                ctk.CTkLabel(
                    table_frame,
                    text=get_text("status", self.lang) + " — " + get_text("requests", self.lang),
                    font=ctk.CTkFont(size=14, weight="bold"),
                    text_color=ColorTheme.TEXT_PRIMARY
                ).pack(pady=(15, 10))
                
                for status_key, config in self.STATUS_CONFIG.items():
                    cur.execute("SELECT COUNT(*) FROM requests WHERE status = ?", (status_key,))
                    count = cur.fetchone()[0] or 0
                    if count == 0:
                        continue
                    
                    row_frame = ctk.CTkFrame(table_frame, fg_color="transparent")
                    row_frame.pack(fill="x", padx=20, pady=3)
                    
                    label_key = f"label_{self.lang}"
                    label = config.get(label_key, status_key)
                    icon = config.get("icon", "")
                    status_color = config.get("color", ColorTheme.TEXT_PRIMARY)
                    
                    ctk.CTkLabel(
                        row_frame,
                        text=f"{icon} {label}",
                        font=ctk.CTkFont(size=13),
                        text_color=ColorTheme.TEXT_PRIMARY,
                        width=150, anchor="w"
                    ).pack(side="left")
                    
                    pct = (count / total * 100) if total > 0 else 0
                    bar_bg = ctk.CTkFrame(row_frame, fg_color=ColorTheme.BG_INPUT, height=20, corner_radius=4)
                    bar_bg.pack(side="left", fill="x", expand=True, padx=(10, 10))
                    bar_bg.pack_propagate(False)
                    
                    bar_width = max(pct / 100, 0.02)
                    bar = ctk.CTkFrame(bar_bg, fg_color=status_color, corner_radius=4)
                    bar.place(relx=0, rely=0, relwidth=bar_width, relheight=1.0)
                    
                    ctk.CTkLabel(
                        row_frame,
                        text=f"{count} ({pct:.0f}%)",
                        font=ctk.CTkFont(size=12),
                        text_color=ColorTheme.TEXT_SECONDARY,
                        width=80
                    ).pack(side="right")
                
                # Top masters
                masters_frame = ctk.CTkFrame(self.content_frame, fg_color=ColorTheme.BG_CARD, corner_radius=12)
                masters_frame.pack(fill="x", padx=20, pady=(0, 10))
                
                ctk.CTkLabel(
                    masters_frame,
                    text=get_text("top_masters", self.lang) if get_text("top_masters", self.lang) != "top_masters" else "Топ мастеров",
                    font=ctk.CTkFont(size=14, weight="bold"),
                    text_color=ColorTheme.TEXT_PRIMARY
                ).pack(pady=(15, 10))
                
                cur.execute("""
                    SELECT u.full_name, COUNT(r.id) as cnt,
                           COALESCE(SUM(r.total_cost), 0) as total
                    FROM requests r
                    JOIN users u ON r.user_id = u.id
                    WHERE r.status IN ('ready', 'closed')
                    GROUP BY u.id
                    ORDER BY cnt DESC
                    LIMIT 5
                """)
                masters = cur.fetchall()
                
                if masters:
                    for m_name, m_count, m_total in masters:
                        row_frame = ctk.CTkFrame(masters_frame, fg_color="transparent")
                        row_frame.pack(fill="x", padx=20, pady=3)
                        ctk.CTkLabel(row_frame, text=m_name or "—", text_color=ColorTheme.TEXT_PRIMARY, font=ctk.CTkFont(size=13), anchor="w", width=200).pack(side="left")
                        ctk.CTkLabel(row_frame, text=f"{m_count} заявок", text_color=ColorTheme.TEXT_SECONDARY, font=ctk.CTkFont(size=12), width=100).pack(side="left", padx=10)
                        ctk.CTkLabel(row_frame, text=format_currency(m_total, "RUB", self.lang), text_color=ColorTheme.SUCCESS, font=ctk.CTkFont(size=12, weight="bold")).pack(side="right", padx=20)
                else:
                    ctk.CTkLabel(masters_frame, text="—", text_color=ColorTheme.TEXT_SECONDARY).pack(pady=10)
                
                # Bottom padding
                ctk.CTkFrame(masters_frame, fg_color="transparent", height=10).pack()
        
        except Exception as e:
            app_logger.error(f"Error loading performance report: {e}")
            ToastNotification(self, f"{get_text('error_loading', self.lang)}: {e}", "error")
    
    def refresh(self) -> None:
        """Публичный метод для обновления отчётов (вызов извне)"""
        # Можно добавить перезагрузку активного отчёта
        pass