# services/report_service.py
"""
Расширенный сервис отчётов и статистики для PC Repair CRM Pro

✅ Автоматическая генерация статистики для дашборда
✅ Финансовые отчёты с фильтрацией по периоду
✅ Отчёты по запасам, сотрудникам, контрагентам
✅ Экспорт в CSV, Excel, JSON с валидацией
✅ Кэширование частых запросов для производительности
"""

import sqlite3
import csv
import os
import json
from typing import Dict, List, Optional, Tuple, TypedDict, Any
from datetime import datetime, timedelta
from functools import lru_cache
import re

from database.connection import DatabaseConnection
from core.logger import app_logger
from utils.validators import validate_date_format


# ==================== 🎯 ТИПЫ ДАННЫХ ====================

class DashboardStatsDict(TypedDict):
    """Тип для статистики дашборда"""
    total_requests: int
    active_requests: int
    completed_requests: int
    total_costs: float
    low_stock_parts: int
    new_employees_this_month: int  # ✅ employees, не clients!
    revenue_this_month: float


class FinancialReportDict(TypedDict):
    """Тип для финансового отчёта"""
    total_requests: int
    revenue: float
    labor_total: float
    parts_total: float
    avg_check: float
    status_breakdown: Dict[str, Dict[str, Any]]


class InventoryReportDict(TypedDict):
    """Тип для отчёта по запасам"""
    total_items: int
    total_quantity: int
    total_cost_value: float
    total_retail_value: float
    top_parts: List[Tuple]
    low_stock: int
    out_of_stock: int


# ✅ ИМПОРТ МОДЕЛИ DashboardStats (если есть)
try:
    from models.report import DashboardStats
except ImportError:
    # Создаём простую замену если модель не найдена
    from dataclasses import dataclass

    @dataclass
    class DashboardStats:
        total_requests: int = 0
        active_requests: int = 0
        completed_requests: int = 0
        total_costs: float = 0.0
        low_stock_parts: int = 0
        new_employees_this_month: int = 0  # ✅ employees, не clients!
        revenue_this_month: float = 0.0

        def to_dict(self) -> DashboardStatsDict:
            return {
                "total_requests": self.total_requests,
                "active": self.active_requests,
                "completed": self.completed_requests,
                "total_costs": self.total_costs,
                "low_stock": self.low_stock_parts,
                "new_employees": self.new_employees_this_month,
                "revenue": self.revenue_this_month,
            }


# ==================== ⚙️ КОНФИГУРАЦИЯ ====================

# ✅ Whitelist разрешённых таблиц для экспорта (защита от SQL injection)
ALLOWED_EXPORT_TABLES = {
    "requests", "employees", "contractors", "parts", 
    "equipment", "directories", "users", "branches"
}

# ✅ Пороги для отчётов
LOW_STOCK_THRESHOLD = 5
CACHE_TTL_SECONDS = 300  # 5 минут кэша для dashboard stats


class ReportService:
    """
    Сервис генерации отчётов и статистики
    
    ✅ Кэширование частых запросов через @lru_cache
    ✅ Валидация входных данных (даты, названия таблиц)
    ✅ Поддержка TypedDict для типобезопасности
    ✅ Экспорт в несколько форматов с обработкой ошибок
    ✅ Корректные запросы к таблице employees (не clients)
    
    Пример использования:
        >>> service = ReportService()
        >>> stats = service.get_dashboard_stats()
        >>> print(f"Активных заявок: {stats.active_requests}")
        
        >>> financial = service.get_financial_report("2024-01-01", "2024-01-31")
        >>> print(f"Выручка: {financial['revenue']:.2f} ₽")
    """
    
    def __init__(self, db: Optional[DatabaseConnection] = None):
        self.db = db or DatabaseConnection()
        self._stats_cache: Optional[Tuple[DashboardStats, float]] = None  # (stats, timestamp)
        
        app_logger.info("📊 ReportService initialized")
    
    def _is_cache_valid(self) -> bool:
        """Проверка актуальности кэша"""
        if self._stats_cache is None:
            return False
        _, timestamp = self._stats_cache
        return (datetime.now().timestamp() - timestamp) < CACHE_TTL_SECONDS
    
    # ==================== 🎯 DASHBOARD STATS ====================
    
    @lru_cache(maxsize=1)
    def get_dashboard_stats(self) -> DashboardStats:
        """
        Получение статистики для дашборда с кэшированием
        
        ✅ Кэш на 5 минут для снижения нагрузки на БД
        ✅ Корректные запросы к таблице employees (не clients)
        
        Returns:
            DashboardStats: Объект со статистикой
        """
        # ✅ Проверяем кэш
        if self._is_cache_valid():
            app_logger.debug("📊 Using cached dashboard stats")
            return self._stats_cache[0]
        
        try:
            with self.db.get_cursor() as cur:
                # Статистика по заявкам
                cur.execute("""
                    SELECT 
                        COUNT(*) as total,
                        SUM(CASE WHEN status IN ('new', 'diagnostics', 'in_progress') THEN 1 ELSE 0 END) as active,
                        SUM(CASE WHEN status = 'closed' THEN 1 ELSE 0 END) as completed,
                        COALESCE(SUM(total_cost), 0) as total_costs
                    FROM requests
                """)
                req_row = cur.fetchone() or (0, 0, 0, 0)

                # Низкий остаток запчастей
                cur.execute(
                    f"SELECT COUNT(*) FROM parts WHERE quantity <= ? AND quantity > 0",
                    (LOW_STOCK_THRESHOLD,)
                )
                low_stock = cur.fetchone()[0] or 0

                # ✅ Новые сотрудники за месяц (не клиенты!)
                cur.execute("""
                    SELECT COUNT(*) FROM employees 
                    WHERE created_at >= date('now', '-1 month')
                """)
                new_employees = cur.fetchone()[0] or 0

                # Выручка за месяц
                cur.execute("""
                    SELECT COALESCE(SUM(total_cost), 0) FROM requests
                    WHERE created_at >= date('now', '-1 month')
                    AND status = 'closed'
                """)
                monthly_revenue = cur.fetchone()[0] or 0

                stats = DashboardStats(
                    total_requests=req_row[0] or 0,
                    active_requests=req_row[1] or 0,
                    completed_requests=req_row[2] or 0,
                    total_costs=req_row[3] or 0,
                    low_stock_parts=low_stock,
                    new_employees_this_month=new_employees,  # ✅ employees, не clients
                    revenue_this_month=monthly_revenue,
                )
                
                # ✅ Сохраняем в кэш
                self._stats_cache = (stats, datetime.now().timestamp())
                
                return stats

        except Exception as e:
            app_logger.exception(f"❌ Error getting dashboard stats: {e}")
            return DashboardStats()
    
    def clear_stats_cache(self) -> None:
        """Очистить кэш статистики (вызывать после изменений в БД)"""
        self._stats_cache = None
        if hasattr(self.get_dashboard_stats, 'cache_clear'):
            self.get_dashboard_stats.cache_clear()
    
    # ==================== 📊 ФИНАНСОВЫЕ ОТЧЁТЫ ====================
    
    def get_financial_report(self, start_date: str, end_date: str) -> FinancialReportDict:
        """
        Финансовый отчёт за период с валидацией дат
        
        ✅ Проверка формата дат "%Y-%m-%d"
        ✅ Проверка логики: start_date <= end_date
        
        Args:
            start_date: Дата начала в формате "YYYY-MM-DD"
            end_date: Дата окончания в формате "YYYY-MM-DD"
            
        Returns:
            FinancialReportDict: Словарь с финансовыми метриками
        """
        # ✅ Валидация формата дат
        if not validate_date_format(start_date) or not validate_date_format(end_date):
            app_logger.error(f"❌ Invalid date format: {start_date} or {end_date}")
            return {}
        
        # ✅ Валидация логики дат
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            if start_dt > end_dt:
                app_logger.error(f"❌ start_date ({start_date}) > end_date ({end_date})")
                return {}
        except ValueError as e:
            app_logger.error(f"❌ Date parsing error: {e}")
            return {}
        
        try:
            with self.db.get_cursor() as cur:
                cur.execute(
                    """
                    SELECT COUNT(*) as total, 
                           SUM(total_cost) as revenue,
                           SUM(labor_cost) as labor_total,
                           SUM(parts_cost) as parts_total
                    FROM requests 
                    WHERE status = 'closed' 
                    AND DATE(created_at) BETWEEN ? AND ?
                """,
                    (start_date, end_date),
                )
                row = cur.fetchone() or (0, 0, 0, 0)

                total = row[0] or 0
                revenue = row[1] or 0
                avg_check = revenue / total if total > 0 else 0

                cur.execute(
                    """
                    SELECT status, COUNT(*), SUM(total_cost)
                    FROM requests
                    WHERE DATE(created_at) BETWEEN ? AND ?
                    GROUP BY status
                """,
                    (start_date, end_date),
                )
                status_breakdown = {
                    r[0]: {"count": r[1] or 0, "sum": r[2] or 0} 
                    for r in cur.fetchall()
                }

                return {
                    "total_requests": total,
                    "revenue": revenue,
                    "labor_total": row[2] or 0,
                    "parts_total": row[3] or 0,
                    "avg_check": round(avg_check, 2),
                    "status_breakdown": status_breakdown,
                }
        except Exception as e:
            app_logger.exception(f"❌ Financial report error: {e}")
            return {}
    
    def get_inventory_report(self) -> InventoryReportDict:
        """Отчёт по запасам с параметризованным порогом"""
        try:
            with self.db.get_cursor() as cur:
                cur.execute(
                    "SELECT COUNT(*), SUM(quantity), SUM(quantity * cost), SUM(quantity * price) FROM parts"
                )
                row = cur.fetchone() or (0, 0, 0, 0)

                cur.execute("""
                    SELECT name, quantity, cost, price, (quantity * cost) as total_value
                    FROM parts ORDER BY total_value DESC LIMIT 10
                """)
                top_parts = cur.fetchall()

                cur.execute(
                    f"SELECT COUNT(*) FROM parts WHERE quantity <= min_stock AND quantity > 0"
                )
                low_stock = cur.fetchone()[0] or 0

                cur.execute("SELECT COUNT(*) FROM parts WHERE quantity = 0")
                out_of_stock = cur.fetchone()[0] or 0

                return {
                    "total_items": row[0] or 0,
                    "total_quantity": row[1] or 0,
                    "total_cost_value": row[2] or 0,
                    "total_retail_value": row[3] or 0,
                    "top_parts": top_parts,
                    "low_stock": low_stock,
                    "out_of_stock": out_of_stock,
                }
        except Exception as e:
            app_logger.exception(f"❌ Inventory report error: {e}")
            return {}
    
    def get_employee_statistics(self) -> Dict[str, Any]:  # ✅ employees, не clients
        """Статистика по сотрудникам (бывш. клиентам)"""
        try:
            with self.db.get_cursor() as cur:
                # ✅ employees вместо clients
                cur.execute("SELECT COUNT(*) FROM employees")
                total_employees = cur.fetchone()[0] or 0

                cur.execute("""
                    SELECT e.full_name, COUNT(r.id) as requests_count, SUM(r.total_cost) as total_spent
                    FROM employees e
                    LEFT JOIN requests r ON e.id = r.client_id
                    WHERE r.status = 'closed'
                    GROUP BY e.id ORDER BY total_spent DESC LIMIT 10
                """)
                top_employees = cur.fetchall()

                cur.execute(
                    "SELECT COUNT(*) FROM employees WHERE created_at >= date('now', '-1 month')"
                )
                new_employees_month = cur.fetchone()[0] or 0

                return {
                    "total_employees": total_employees,
                    "top_employees": top_employees,
                    "new_employees_month": new_employees_month,
                }
        except Exception as e:
            app_logger.exception(f"❌ Employee stats error: {e}")
            return {}
    
    # ==================== 📤 ЭКСПОРТ ====================
    
    def export_to_csv(self, table_name: str, output_path: str) -> bool:
        """
        Экспорт таблицы в CSV с валидацией имени таблицы
        
        ✅ Защита от SQL injection через whitelist
        ✅ Экранирование специальных символов в CSV
        
        Args:
            table_name: Имя таблицы (должно быть в ALLOWED_EXPORT_TABLES)
            output_path: Путь для сохранения файла
            
        Returns:
            bool: True если экспорт успешен
        """
        # ✅ Валидация названия таблицы
        if table_name not in ALLOWED_EXPORT_TABLES:
            app_logger.error(f"❌ Invalid table name for export: {table_name}")
            return False
        
        try:
            with self.db.get_cursor() as cur:
                # ✅ Безопасный запрос с валидированным именем таблицы
                cur.execute(f"SELECT * FROM {table_name}")
                rows = cur.fetchall()
                columns = [desc[0] for desc in cur.description] if cur.description else []

            # ✅ Создаём директорию если не существует
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)

            with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
                # ✅ Используем excel диалект для правильного экранирования
                writer = csv.writer(f, dialect='excel', quoting=csv.QUOTE_MINIMAL)
                writer.writerow(columns)
                writer.writerows(rows)

            app_logger.info(f"📤 Exported {table_name} to {output_path} ({len(rows)} rows)")
            return True
        except PermissionError:
            app_logger.error(f"❌ Permission denied: {output_path}")
            return False
        except Exception as e:
            app_logger.exception(f"❌ CSV export error: {e}")
            return False
    
    def export_to_excel(self, data: Dict[str, Any], output_path: str, report_type: str = "financial") -> bool:
        """
        Создание Excel файла с отчётом
        
        ✅ Использование цветов из темы (если доступны)
        ✅ Обработка ImportError для openpyxl
        ✅ Автоматическое создание директории
        
        Args:
            data: Данные отчёта (словарь)
            output_path: Путь для сохранения файла
            report_type: Тип отчёта ("financial", "inventory", "pricelist")
            
        Returns:
            bool: True если экспорт успешен
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # ✅ Создаём директорию если не существует
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)

            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Финансы"

            # ✅ Используем цвета из темы или дефолтные
            try:
                from ui.theme import ColorTheme
                header_color = ColorTheme.PRIMARY
            except ImportError:
                header_color = "6366f1"  # Fallback цвет
            
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(
                start_color=header_color, end_color=header_color, fill_type="solid"
            )
            center_align = Alignment(horizontal="center")

            # Заголовок отчёта
            ws1.merge_cells('A1:C1')
            ws1['A1'] = "📊 ФИНАНСОВЫЙ ОТЧЁТ"
            ws1['A1'].font = Font(bold=True, size=16)
            ws1['A1'].alignment = center_align
            
            ws1.append(["Период:", data.get("period", "Все время"), ""])
            ws1.append([""])
            ws1.append(["Метрика", "Значение", ""])

            # Стили заголовков
            for cell in ws1[4]:
                if cell.value:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align

            # Данные отчёта
            metrics = [
                ("Всего заявок", data.get("total_requests", 0)),
                ("Общая выручка", f"{data.get('revenue', 0):.2f} ₽"),
                ("Средний чек", f"{data.get('avg_check', 0):.2f} ₽"),
                ("Работа", f"{data.get('labor_total', 0):.2f} ₽"),
                ("Запчасти", f"{data.get('parts_total', 0):.2f} ₽"),
            ]
            
            for label, value in metrics:
                ws1.append([label, value, ""])

            # Авто-ширина колонок
            ws1.column_dimensions["A"].width = 25
            ws1.column_dimensions["B"].width = 20
            ws1.column_dimensions["C"].width = 5

            wb.save(output_path)
            app_logger.info(f"📤 Excel report saved to {output_path}")
            return True
        except ImportError:
            app_logger.warning("⚠️ openpyxl not installed, skipping Excel export")
            return False
        except PermissionError:
            app_logger.error(f"❌ Permission denied: {output_path}")
            return False
        except Exception as e:
            app_logger.exception(f"❌ Excel export error: {e}")
            return False
    
    def export_price_list(self, output_path: str) -> bool:
        """Экспорт прайс-листа в Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # ✅ Создаём директорию если не существует
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)

            wb = Workbook()
            ws = wb.active
            ws.title = "Прайс-лист"

            # Стили
            header_font = Font(bold=True, size=14)
            center_align = Alignment(horizontal="center")
            
            ws.merge_cells('A1:F1')
            ws['A1'] = "ПРАЙС-ЛИСТ PC REPAIR CRM PRO"
            ws['A1'].font = header_font
            ws['A1'].alignment = center_align
            
            ws.append([f"Дата: {datetime.now().strftime('%d.%m.%Y')}"])
            ws.append([""])
            ws.append(["№", "Название", "Артикул", "Кол-во", "Цена", "Наличие"])

            with self.db.get_cursor() as cur:
                cur.execute(
                    "SELECT name, sku, quantity, price FROM parts ORDER BY name"
                )
                rows = cur.fetchall()

            for idx, row in enumerate(rows, 1):
                name, sku, quantity, price = row
                status = "✅ В наличии" if (quantity or 0) > 0 else "❌ Нет"
                price_formatted = f"{price or 0:.2f} ₽"
                ws.append([idx, name, sku, quantity or 0, price_formatted, status])

            # Авто-ширина колонок
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

            wb.save(output_path)
            return True
        except ImportError:
            app_logger.warning("⚠️ openpyxl not installed")
            return False
        except Exception as e:
            app_logger.exception(f"❌ Price list export error: {e}")
            return False
    
    def export_to_json(self, report_type: str, data: Dict[str, Any], output_path: str) -> bool:
        """
        Экспорт отчёта в JSON формат
        
        ✅ Поддержка разных типов отчётов
        ✅ Красивое форматирование с отступами
        
        Args:
            report_type: Тип отчёта ("financial", "inventory", "employees")
            data: Данные отчёта
            output_path: Путь для сохранения файла
            
        Returns:
            bool: True если экспорт успешен
        """
        try:
            # ✅ Создаём директорию если не существует
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
            
            export_data = {
                "metadata": {
                    "report_type": report_type,
                    "generated_at": datetime.now().isoformat(),
                    "version": "1.0",
                },
                "data": data,
            }
            
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2, default=str)
            
            app_logger.info(f"📤 JSON report saved to {output_path}")
            return True
        except PermissionError:
            app_logger.error(f"❌ Permission denied: {output_path}")
            return False
        except Exception as e:
            app_logger.exception(f"❌ JSON export error: {e}")
            return False